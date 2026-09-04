#!/usr/bin/env python3
"""Public Google Drive folder synchronizer for the Lucra catalogue."""

from __future__ import annotations

import html
import io
import json
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from decimal import Decimal, ROUND_HALF_UP
from functools import lru_cache
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook

ROOT_FOLDER_ID = "17u1Vo3es5lO07Z0__mfu5ugXCOaTkf4Z"
ROOT_FOLDER_URL = f"https://drive.google.com/drive/folders/{ROOT_FOLDER_ID}"
OUTPUT = Path(__file__).parent / "data" / "inventory.json"
OVERFLOW_MANIFEST_PATH = Path(__file__).parent / "data" / "drive_overflow.json"
USER_AGENT = "Mozilla/5.0 (compatible; LucraInventory/1.0)"
ITEM_PATTERN = re.compile(
    r'<div[^>]*data-id="([^"]+)"[^>]*>.*?<strong[^>]*>(.*?)</strong>', re.S
)
ROW_ITEM_PATTERN = re.compile(r'<tr[^>]*data-id="([^"]+)"[^>]*>(.*?)</tr>', re.S)
STRONG_PATTERN = re.compile(r'<strong[^>]*>(.*?)</strong>', re.S)
ARIA_LABEL_PATTERN = re.compile(r'aria-label="([^"]+)"', re.I)
TAG_PATTERN = re.compile(r"<[^>]+>")
SLAB_PATTERN = re.compile(r"^\s*(\d+)\s*\.\s*(jpe?g|png|webp|heic|heif)$", re.I)
IMAGE_PATTERN = re.compile(r"\.(jpe?g|png|webp|heic|heif)$", re.I)
VIDEO_PATTERN = re.compile(r"\.(mp4|mov|webm)$", re.I)
FILE_PATTERN = re.compile(r"\.(jpe?g|png|webp|heic|heif|mp4|mov|webm|xlsx?)$", re.I)
CAMERA_IMAGE_PATTERN = re.compile(r"^(?:IMG|DSC|PXL)[ _-]?\d+\.(?:jpe?g|png|webp|heic|heif)$", re.I)
PACKING_LIST_PATTERN = re.compile(r"\bpacking\s+list\b", re.I)
REQUEST_LOCK = threading.Lock()
LAST_REQUEST_AT = 0.0
MIN_REQUEST_GAP = 0.35
NESTED_FOLDER_TIMEOUT = 8
NESTED_FOLDER_ATTEMPTS = 1
MAX_WORKERS = 3
L1014_PATTERN = re.compile(r"\bL1014\b", re.I)

try:
    OVERFLOW_MANIFEST = json.loads(OVERFLOW_MANIFEST_PATH.read_text(encoding="utf-8"))
except (FileNotFoundError, json.JSONDecodeError):
    OVERFLOW_MANIFEST = {}


def fetch(url: str, timeout: int = 35, attempts: int = 3) -> bytes:
    global LAST_REQUEST_AT
    last_error = None
    for attempt in range(attempts):
        with REQUEST_LOCK:
            wait = MIN_REQUEST_GAP - (time.monotonic() - LAST_REQUEST_AT)
            if wait > 0:
                time.sleep(wait)
            LAST_REQUEST_AT = time.monotonic()
        request = Request(url, headers={"User-Agent": USER_AGENT})
        try:
            with urlopen(request, timeout=timeout) as response:
                return response.read()
        except Exception as exc:
            last_error = exc
            if attempt < attempts - 1:
                time.sleep(1.5 * (2**attempt))
    raise last_error


def describe_error(exc: Exception) -> str:
    """Return a short, useful error for the sync report and UI."""
    message = str(exc).strip()
    return f"{type(exc).__name__}: {message}" if message else type(exc).__name__


def is_file_item(name: str) -> bool:
    return bool(FILE_PATTERN.search(name) or PACKING_LIST_PATTERN.search(name))


def is_packing_list_item(item: dict[str, str]) -> bool:
    return bool(PACKING_LIST_PATTERN.search(item.get("name", "")) or item.get("kind") == "spreadsheet")


def is_bundle_folder(name: str) -> bool:
    """Recognize coded bundle folders even when their media listing is sparse."""
    clean_name = re.sub(r"^\s*reserved\b\s*(?:-\s*)?", "", name, flags=re.I).strip()
    return bool(re.search(r"\s[KLM]\d+\s*$", clean_name, re.I))


def is_l1014_folder(name: str) -> bool:
    return bool(L1014_PATTERN.search(name))


def inventory_script_path() -> Path:
    return OUTPUT.with_name("inventory.js")


def inventory_script_contents(payload: dict) -> str:
    return "// Generated from inventory.json by sync_drive.py.\nwindow.LUCRA_INVENTORY = " + json.dumps(
        payload, ensure_ascii=False, indent=2
    ) + ";\n"


@lru_cache(maxsize=512)
def folder_items(folder_id: str, timeout: int = 35, attempts: int = 3) -> list[dict[str, str]]:
    raw = fetch(f"https://drive.google.com/drive/folders/{folder_id}", timeout=timeout, attempts=attempts).decode("utf-8")
    found: list[dict[str, str]] = []
    seen: set[str] = set()

    # In the public Drive list, some image files have no extension. L1014
    # uses names such as "3" and "Bookmatched1". The row markup identifies
    # these as images even though the name alone looks like a folder.
    row_items = []
    for file_id, row in ROW_ITEM_PATTERN.findall(raw):
        names = STRONG_PATTERN.findall(row)
        aria_labels = ARIA_LABEL_PATTERN.findall(row)
        if names:
            raw_name = names[-1]
        else:
            raw_name = next(
                (
                    candidate
                    for label in reversed(aria_labels)
                    for candidate in [
                        re.sub(
                            r"\s+(?:image|folder|file|spreadsheet|microsoft excel|video|shared)\b.*$",
                            "",
                            label,
                            flags=re.I,
                        ).strip()
                    ]
                    if candidate
                ),
                None,
            )
        if not raw_name:
            continue
        if re.search(r"<title>\s*image\s*</title>|aria-label=\"[^\"]*\bimage\b[^\"]*\"", row, re.I):
            kind = "image"
        elif re.search(r"microsoft\s+excel|spreadsheet", row, re.I):
            kind = "spreadsheet"
        elif re.search(r"<title>\s*video\s*</title>|aria-label=\"[^\"]*\bvideo\b[^\"]*\"", row, re.I):
            kind = "video"
        else:
            kind = "unknown"
        row_items.append((file_id, raw_name, kind))
    candidates = row_items or [(file_id, raw_name, "unknown") for file_id, raw_name in ITEM_PATTERN.findall(raw)]
    for file_id, raw_name, kind in candidates:
        if file_id in seen:
            continue
        name = " ".join(html.unescape(TAG_PATTERN.sub("", raw_name)).split())
        found.append({"id": file_id, "name": name, "kind": kind})
        seen.add(file_id)
    # Google’s anonymous folder HTML currently exposes only the first 50
    # children for larger public folders. This local, read-only continuation
    # cache contains the remaining public file IDs discovered during audit;
    # the files themselves are still downloaded live from Drive during sync.
    for item in OVERFLOW_MANIFEST.get(folder_id, {}).get("items", []):
        if item["id"] in seen:
            continue
        found.append({"id": item["id"], "name": item["name"], "kind": item.get("kind", "unknown")})
        seen.add(item["id"])
    return found


def download_file(file_id: str) -> bytes:
    query = urlencode({"id": file_id, "export": "download", "confirm": "t"})
    return fetch(f"https://drive.usercontent.google.com/download?{query}", timeout=45)


def clean_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, str):
        value = value.strip().replace("\xa0", " ").replace(",", ".")
        match = re.search(r"-?\d+(?:\.\d+)?", value)
        if not match:
            return value
        value = match.group(0)
    try:
        number = float(value)
        return int(number) if number.is_integer() else round(number, 4)
    except (TypeError, ValueError):
        return value


def _header_text(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _numeric(value):
    if isinstance(value, str) and not re.fullmatch(r"\s*-?\d+(?:[.,]\d+)?\s*", value):
        return None
    number = clean_number(value)
    return number if isinstance(number, (int, float)) and not isinstance(number, bool) else None


def _round_area(value) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _find_column(headers: list[str], predicate) -> int | None:
    return next((index for index, header in enumerate(headers) if predicate(header)), None)


def _packing_mapping(rows: list[list]) -> dict | None:
    """Find common packing-list layouts, including merged/multi-row headers."""
    for header_index, row in enumerate(rows):
        headers = [_header_text(cell) for cell in row]
        has_block = any("block" in header or "bundle" in header for header in headers)
        has_measure = any("pcs" in header or "sqm" in header or "m2" in header for header in headers)
        if not has_block or not has_measure:
            continue

        block_col = _find_column(headers, lambda header: "block" in header or "bundle" in header)
        material_col = _find_column(headers, lambda header: any(token in header for token in ("material", "product", "description")))
        finish_col = _find_column(headers, lambda header: "finish" in header or "surface" in header)
        pcs_col = _find_column(headers, lambda header: "pcs" in header or "pieces" in header)
        sqm_candidates = [
            index
            for index, header in enumerate(headers)
            if "sqm" in header or "m2" in header or "m²" in header
        ]
        sqm_col = next(
            (
                index
                for index in sqm_candidates
                if not any(token in headers[index] for token in ("total", "bundle", "bandil", "band l"))
            ),
            sqm_candidates[0] if sqm_candidates else None,
        )
        width_col = _find_column(headers, lambda header: header in {"width", "w"} or header.startswith("width "))
        height_col = _find_column(headers, lambda header: header in {"height", "h"} or header.startswith("height "))
        length_col = _find_column(headers, lambda header: header in {"length", "lenght", "l"} or header.startswith("length ") or header.startswith("lenght "))

        # Some workbooks put T/W/L on the row below a merged NET SIZES header.
        for extra_row in rows[header_index + 1 : header_index + 4]:
            for index, cell in enumerate(extra_row):
                label = _header_text(cell)
                if label == "w" and width_col is None:
                    width_col = index
                elif label == "l" and length_col is None:
                    length_col = index
        if width_col is None and height_col is not None and length_col is not None:
            width_col, height_col = height_col, length_col
        elif width_col is None and length_col is not None:
            width_col = length_col
        if height_col is None and length_col is not None:
            height_col = length_col

        if block_col is not None and material_col is not None and pcs_col is not None and sqm_col is not None:
            # A merged "Material Name" header is used by some lists even
            # though each data row stores finish first and material second.
            # Likewise, a merged "Dimensions" header may leave W/L unnamed.
            sample = next(
                (
                    candidate
                    for candidate in rows[header_index + 1 : header_index + 8]
                    if any(str(cell or "").strip() for cell in candidate[block_col:])
                    and len([index for index in range(block_col + 1, len(candidate)) if _numeric(candidate[index]) is not None]) >= 4
                ),
                None,
            )
            if sample is not None:
                numeric_cols = [index for index in range(block_col + 1, len(sample)) if _numeric(sample[index]) is not None]
                if width_col is None or height_col is None:
                    dimension_cols = [index for index in numeric_cols if index not in {pcs_col, sqm_col}]
                    if width_col is None and dimension_cols:
                        width_col = dimension_cols[0]
                    if height_col is None and len(dimension_cols) > 1:
                        height_col = dimension_cols[1]
                if finish_col is None:
                    text_cols = [
                        index
                        for index in range(block_col + 1, numeric_cols[0])
                        if str(sample[index] or "").strip()
                    ]
                    if len(text_cols) >= 2:
                        first_text, second_text = (str(sample[index]).strip() for index in text_cols[:2])
                        first_is_finish = bool(re.search(r"polish|honed|raw|leather|surface|finish", first_text, re.I))
                        second_is_thickness = bool(re.fullmatch(r"\d+(?:[.,]\d+)?\s*(?:cm|mm)", second_text, re.I))
                        if second_is_thickness:
                            # The second text cell is thickness, not a
                            # finish; the product cell may already include
                            # its finish wording.
                            material_col, finish_col = text_cols[0], None
                        elif first_is_finish:
                            finish_col, material_col = text_cols[0], text_cols[1]
                        else:
                            material_col, finish_col = text_cols[0], text_cols[1]
            return {
                "start": header_index + 1,
                "block": block_col,
                "material": material_col,
                "finish": finish_col,
                "width": width_col,
                "height": height_col,
                "pcs": pcs_col,
                "sqm": sqm_col,
            }

    # A few packing lists have no labels at all, but their first data row has
    # a block reference followed by finish/material text and four numeric
    # columns: width, length, pcs, sqm.
    for row_index, row in enumerate(rows):
        block_col = next(
            (
                index
                for index, cell in enumerate(row)
                if re.fullmatch(r"[A-Z]{1,3}[0-9][A-Z0-9-]{3,}", str(cell or "").strip(), re.I)
            ),
            None,
        )
        if block_col is None:
            continue
        numeric_cols = [index for index in range(block_col + 1, len(row)) if _numeric(row[index]) is not None]
        if len(numeric_cols) < 4:
            continue
        text_cols = [index for index in range(block_col + 1, numeric_cols[0]) if str(row[index] or "").strip()]
        if len(text_cols) < 2:
            continue
        first_text, second_text = str(row[text_cols[0]]).strip(), str(row[text_cols[1]]).strip()
        finish_first = bool(re.search(r"polish|honed|raw|leather|surface|finish", first_text, re.I))
        return {
            "start": row_index,
            "block": block_col,
            "material": text_cols[1] if finish_first else text_cols[0],
            "finish": text_cols[0] if finish_first else text_cols[1],
            "width": numeric_cols[0],
            "height": numeric_cols[1],
            "pcs": numeric_cols[2],
            "sqm": numeric_cols[3],
        }
    return None


def parse_packing_list(content: bytes) -> dict:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    rows = [list(row) for row in workbook.active.iter_rows(values_only=True)]
    mapping = _packing_mapping(rows)
    if mapping is None:
        return {"lines": [], "totalPcs": None, "totalSqm": None}

    lines = []
    current_material = ""
    current_finish = ""
    for row in rows[mapping["start"] :]:
        cells = list(row)

        def cell_at(column):
            return cells[column] if column is not None and column < len(cells) else None

        raw_block = cell_at(mapping["block"])
        raw_material = cell_at(mapping["material"])
        raw_finish = cell_at(mapping["finish"])
        if str(raw_material or "").strip():
            current_material = str(raw_material).strip()
        if str(raw_finish or "").strip():
            current_finish = str(raw_finish).strip()
        material = current_material
        finish = current_finish
        width, height = _numeric(cell_at(mapping["width"])), _numeric(cell_at(mapping["height"]))
        pcs, sqm = _numeric(cell_at(mapping["pcs"])), _numeric(cell_at(mapping["sqm"]))
        if not material or pcs is None:
            continue
        # Skip grand-total rows that would otherwise inherit the previous
        # material. Continuation rows still pass because they have dimensions.
        if width is None and height is None and not str(raw_block or "").strip() and not str(raw_material or "").strip():
            continue
        if sqm is None and width is not None and height is not None:
            sqm = _round_area(width * height * pcs / 10000)
        lines.append(
            {
                "block": str(raw_block or "").strip(),
                "finish": finish,
                "material": material,
                "widthCm": width,
                "heightCm": height,
                "pcs": pcs,
                "sqm": sqm,
            }
        )
    total_pcs = sum(line["pcs"] for line in lines if isinstance(line["pcs"], (int, float)))
    total_sqm = _round_area(sum(line["sqm"] for line in lines if isinstance(line["sqm"], (int, float))))
    return {"lines": lines, "totalPcs": total_pcs or None, "totalSqm": total_sqm or None}


def collect_media(items: list[dict[str, str]]) -> tuple[list, list, list, dict | None, str | None, list[dict[str, str]]]:
    slab_images, extra_images, videos = [], [], []
    camera_images = []
    nested_folders = []
    skipped_photo_folders = []
    packing = None
    packing_name = None

    def read_packing(item):
        nonlocal packing, packing_name
        if packing_name:
            return
        packing_name = item["name"]
        try:
            packing = parse_packing_list(download_file(item["id"]))
        except Exception as exc:
            packing = {"lines": [], "totalPcs": None, "totalSqm": None, "error": describe_error(exc)}

    def read_file(item, *, nested_folder_name=None):
        name = item["name"]
        slab_match = SLAB_PATTERN.match(name)
        if slab_match and (not nested_folder_name or nested_folder_name.isdigit()):
            number = int(slab_match.group(1))
            if nested_folder_name and nested_folder_name.isdigit():
                number = int(nested_folder_name)
                slab_images.append({"number": number, "label": str(number), "view": len([image for image in slab_images if image["number"] == number]) + 1, "fileId": item["id"], "name": name})
            else:
                slab_images.append({"number": number, "label": str(number), "fileId": item["id"], "name": name})
            return
        if IMAGE_PATTERN.search(name):
            if CAMERA_IMAGE_PATTERN.match(name) and not nested_folder_name:
                camera_images.append(item)
            elif nested_folder_name and nested_folder_name.isdigit():
                number = int(nested_folder_name)
                slab_images.append({"number": number, "label": str(number), "view": len([image for image in slab_images if image["number"] == number]) + 1, "fileId": item["id"], "name": name})
            else:
                label = re.sub(r"\s*\.\s*(jpe?g|png|webp|heic|heif)$", "", name, flags=re.I).strip()
                extra_images.append({"label": label, "fileId": item["id"], "name": name, "compact": bool(nested_folder_name)})
            return
        if item.get("kind") == "image":
            label = name.strip()
            if nested_folder_name and nested_folder_name.isdigit():
                number = int(nested_folder_name)
                slab_images.append({"number": number, "label": str(number), "view": len([image for image in slab_images if image["number"] == number]) + 1, "fileId": item["id"], "name": name})
            elif label.isdigit():
                slab_images.append({"number": int(label), "label": label, "fileId": item["id"], "name": name})
            else:
                extra_images.append({"label": label, "fileId": item["id"], "name": name, "compact": bool(nested_folder_name)})
            return
        if VIDEO_PATTERN.search(name):
            prefix = f"{nested_folder_name} · " if nested_folder_name else ""
            videos.append({"label": f'{prefix}{name.rsplit(".", 1)[0].strip()}', "fileId": item["id"], "name": name})
            return
        if is_packing_list_item(item):
            read_packing(item)
            return
        if not FILE_PATTERN.search(name):
            nested_folders.append(item)

    for item in items:
        read_file(item)

    def read_nested(item):
        try:
            # Nested photo folders are helpful but should never hold up the
            # bundle itself. Some public Drive subfolders respond slowly or
            # are not independently listable, so use one short attempt here.
            return item, folder_items(item["id"], timeout=NESTED_FOLDER_TIMEOUT, attempts=NESTED_FOLDER_ATTEMPTS), None
        except Exception as exc:
            return item, [], describe_error(exc)

    def walk_nested(item, nested, depth=0):
        folder_name = item["name"].strip()
        nested_folders_inside = []
        for child in nested:
            if IMAGE_PATTERN.search(child["name"]) or child.get("kind") == "image" or VIDEO_PATTERN.search(child["name"]):
                read_file(child, nested_folder_name=folder_name)
            elif is_packing_list_item(child):
                read_packing(child)
            elif not FILE_PATTERN.search(child["name"]):
                nested_folders_inside.append(child)
        if depth >= 2:
            return
        for child_folder in nested_folders_inside:
            try:
                deeper = folder_items(child_folder["id"], timeout=NESTED_FOLDER_TIMEOUT, attempts=NESTED_FOLDER_ATTEMPTS)
                walk_nested(child_folder, deeper, depth + 1)
            except Exception as exc:
                skipped_photo_folders.append(
                    {
                        "name": f'{folder_name} / {child_folder["name"]}',
                        "folderId": child_folder["id"],
                        "error": describe_error(exc),
                    }
                )

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as nested_pool:
        futures = {nested_pool.submit(read_nested, item): item for item in nested_folders}
        for future in as_completed(futures):
            item = futures[future]
            try:
                item, nested, error = future.result()
            except Exception as exc:
                nested, error = [], describe_error(exc)
            if error:
                skipped_photo_folders.append({"name": item["name"], "folderId": item["id"], "error": error})
                continue
            walk_nested(item, nested)
    # Some bundles (notably K6222) contain only camera-style filenames such as
    # IMG_9053.HEIC rather than 1.jpg, 2.jpg, etc. Treat those as sequential
    # slab photos when no numbered slab photos were found, while retaining all
    # other named images as supplementary views.
    if not slab_images and camera_images:
        for number, item in enumerate(sorted(camera_images, key=lambda image: image["name"].lower()), start=1):
            slab_images.append({"number": number, "label": str(number), "fileId": item["id"], "name": item["name"]})
    # Extra-only nested folders are presented as a single numbered sequence so
    # the jump controls stay compact. Numeric slab filenames sort naturally;
    # named views follow them.
    compact_extras = [image for image in extra_images if image.pop("compact", False)]
    compact_extras.sort(
        key=lambda image: (
            0,
            int(re.match(r"^\s*(\d+)", image["name"]).group(1)),
        )
        if re.match(r"^\s*(\d+)", image["name"])
        else (1, image["name"].lower())
    )
    for index, image in enumerate(compact_extras, start=1):
        image["label"] = str(index)
    extra_images.sort(
        key=lambda image: (
            0 if str(image.get("label", "")).isdigit() else 1,
            int(image["label"]) if str(image.get("label", "")).isdigit() else str(image.get("label", "")).lower(),
        )
    )
    skipped_photo_folders.sort(key=lambda folder: (folder["name"].lower(), folder["folderId"]))
    return slab_images, extra_images, videos, packing, packing_name, skipped_photo_folders


def normalize_folder(folder: dict[str, str]) -> dict:
    folder_name = folder["name"]
    reserved = bool(re.match(r"^\s*reserved\b", folder_name, re.I))
    clean_name = re.sub(r"^\s*reserved\b\s*(?:-\s*)?", "", folder_name, flags=re.I).strip()
    code_match = re.search(r"\s([KLM]\d+)\s*$", clean_name, re.I)
    code = code_match.group(1).upper() if code_match else "—"
    display_name = clean_name[: code_match.start()].strip() if code_match else clean_name
    if "_items" in folder:
        items = folder["_items"]
    else:
        # A failed discovery request is retried here. This is especially
        # important for L1014: its bundle must not disappear just because the
        # first folder listing was temporarily slow.
        items = folder_items(folder["id"], timeout=35, attempts=3)
    slab_images, extra_images, videos, packing, packing_name, skipped_photo_folders = collect_media(items)
    slab_images.sort(key=lambda image: (image["number"], image.get("view", 0)))
    packing = packing or {"lines": [], "totalPcs": len(slab_images) or None, "totalSqm": None}
    if code == "—":
        code_sources = [packing_name or ""] + [str(line.get("block") or "") for line in packing.get("lines", [])]
        nested_code = next((match.group(1).upper() for source in code_sources if (match := re.search(r"\b([KLM]\d+)\b", source, re.I))), None)
        if nested_code:
            code = nested_code
    finishes = sorted({line["finish"] for line in packing["lines"] if line.get("finish")})
    dimensions = sorted({f'{line["widthCm"]} × {line["heightCm"]} cm' for line in packing["lines"] if line.get("widthCm") and line.get("heightCm")})
    return {
        "name": display_name,
        "folderName": folder_name,
        "folderId": folder["id"],
        "groupName": folder.get("groupName"),
        "code": code,
        "reserved": reserved,
        "finish": " / ".join(finishes) or "Natural stone",
        "pcs": packing.get("totalPcs") or len(slab_images) or None,
        "sqm": packing.get("totalSqm"),
        "dimensions": dimensions,
        "lines": packing.get("lines", []),
        "images": slab_images,
        "extraImages": extra_images,
        "videos": videos,
        "packingList": packing_name,
        "syncError": packing.get("error"),
        "skippedPhotoFolders": skipped_photo_folders,
    }


def sync_inventory(root_folder_id: str = ROOT_FOLDER_ID) -> dict:
    # folder_items is cached within one sync so repeated lookups are cheap,
    # but every manual sync must begin with a fresh view of public Drive.
    cache_clear = getattr(folder_items, "cache_clear", None)
    if callable(cache_clear):
        cache_clear()
    root_folders = folder_items(root_folder_id)
    folders = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as discovery_pool:
        discovered = {discovery_pool.submit(folder_items, folder["id"]): folder for folder in root_folders}
        for future in as_completed(discovered):
            folder = discovered[future]
            try:
                children = future.result()
            except Exception:
                # Keep the folder as a candidate. normalize_folder retries
                # the listing and the final error can then be reported with
                # the actual bundle name instead of silently treating it as a
                # group with no children.
                folders.append(folder)
                continue
            if is_bundle_folder(folder["name"]) or any(is_file_item(child["name"]) for child in children):
                folders.append({**folder, "_items": children})
            else:
                # Group folders (for example Sunset Dolomite New and Tundra
                # Grey) contain the actual bundles one level below. Preserve
                # the group name on each child so the UI can search/display
                # the family even though the child folder is the product.
                if children:
                    for child in children:
                        child["groupName"] = folder["name"]
                    folders.extend(children)
                else:
                    folders.append({**folder, "_items": children})
    products = []
    errors = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(normalize_folder, folder): folder for folder in folders}
        for future in as_completed(futures):
            folder = futures[future]
            try:
                products.append(future.result())
            except Exception as exc:
                errors.append({"folder": folder["name"], "kind": "bundle", "error": describe_error(exc)})
    products.sort(key=lambda product: (product["name"].lower(), product["code"]))
    warnings = []
    for product in products:
        for skipped in product.get("skippedPhotoFolders", []):
            warnings.append(
                {
                    "folder": product["folderName"],
                    "photoFolder": skipped["name"],
                    "kind": "optional-photo-folder",
                    "error": skipped["error"],
                }
            )
        if product.get("syncError"):
            warnings.append(
                {
                    "folder": product["folderName"],
                    "kind": "packing-list",
                    "error": product["syncError"],
                }
            )
    payload = {
        "source": ROOT_FOLDER_URL,
        "syncedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "location": "Denizli, Türkiye",
        "products": products,
        "errors": errors,
        "warnings": warnings,
    }
    OUTPUT.parent.mkdir(exist_ok=True)
    if errors and OUTPUT.exists():
        previous = json.loads(OUTPUT.read_text(encoding="utf-8"))
        failed_names = {error["folder"] for error in errors}
        products.extend(product for product in previous.get("products", []) if product.get("folderName") in failed_names)
        products.sort(key=lambda product: (product["name"].lower(), product["code"]))
        payload["products"] = products
    # L1014 is a real bundle, not an optional nested photo folder. If its
    # source folder was visible at the root but the bundle could not be
    # normalized, retain the previous record when possible; otherwise fail
    # the sync rather than publishing a catalogue that silently omits it.
    l1014_folders = [folder for folder in root_folders if is_l1014_folder(folder["name"])]
    if l1014_folders and not any(
        any(product.get("folderId") == folder["id"] for folder in l1014_folders) for product in products
    ):
        previous = json.loads(OUTPUT.read_text(encoding="utf-8")) if OUTPUT.exists() else {}
        previous_l1014 = next(
            (
                product
                for product in previous.get("products", [])
                if any(product.get("folderId") == folder["id"] for folder in l1014_folders)
            ),
            None,
        )
        if previous_l1014:
            products.append(previous_l1014)
            products.sort(key=lambda product: (product["name"].lower(), product["code"]))
            payload["products"] = products
            payload["warnings"].append(
                {
                    "folder": l1014_folders[0]["name"],
                    "kind": "required-bundle-retained",
                    "error": "Fresh listing failed; the previous L1014 record was retained.",
                }
            )
        else:
            raise RuntimeError("Rosso Levanto L1014 could not be read; the previous catalogue was preserved.")
    if not products:
        raise RuntimeError("Sync returned no products; the previous catalogue was preserved.")
    temporary = OUTPUT.with_suffix(".json.tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    script_output = inventory_script_path()
    script_temporary = script_output.with_suffix(".js.tmp")
    script_temporary.write_text(inventory_script_contents(payload), encoding="utf-8")
    temporary.replace(OUTPUT)
    script_temporary.replace(script_output)
    return payload


if __name__ == "__main__":
    result = sync_inventory()
    print(f'Synced {len(result["products"])} bundles with {len(result["errors"])} folder errors.')
